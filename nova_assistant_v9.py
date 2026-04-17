"""
=====================================================
  NOVA - AI-POWERED VOICE ASSISTANT 
  By: yash pawar,vaibhav bhawnani,satakshi chaturvedi,tanu kumari
  AI powered by Groq (Free & Fast!)
=====================================================

INSTALL ALL DEPENDENCIES:
pip install SpeechRecognition pyttsx3 pyautogui groq openpyxl
pip install requests pywhatkit deep-translator psutil screen-brightness-control
"""

import os, sys, time, glob, datetime, webbrowser, subprocess, threading, math, re
import random, string

try:
    import speech_recognition as sr
except ImportError:
    print("Missing: pip install SpeechRecognition"); sys.exit(1)

try:
    import pyttsx3
except ImportError:
    print("Missing: pip install pyttsx3"); sys.exit(1)

try:
    import pyautogui
except ImportError:
    print("Missing: pip install pyautogui"); sys.exit(1)

try:
    from groq import Groq
    AI_ENABLED = True
except ImportError:
    print("⚠️  AI features disabled. Install with: pip install groq")
    AI_ENABLED = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_ENABLED = True
except ImportError:
    print("⚠️  Excel creation disabled. Install with: pip install openpyxl")
    EXCEL_ENABLED = False

try:
    import requests
    REQUESTS_ENABLED = True
except ImportError:
    print("⚠️  Weather/News disabled. Install with: pip install requests")
    REQUESTS_ENABLED = False

try:
    import pywhatkit
    WHATSAPP_ENABLED = True
except ImportError:
    print("⚠️  WhatsApp disabled. Install with: pip install pywhatkit")
    WHATSAPP_ENABLED = False

try:
    from deep_translator import GoogleTranslator
    TRANSLATOR_ENABLED = True
except ImportError:
    print("⚠️  Translator disabled. Install with: pip install deep-translator")
    TRANSLATOR_ENABLED = False

try:
    import psutil
    PSUTIL_ENABLED = True
except ImportError:
    print("⚠️  App closer disabled. Install with: pip install psutil")
    PSUTIL_ENABLED = False

try:
    import screen_brightness_control as sbc
    BRIGHTNESS_ENABLED = True
except ImportError:
    print("⚠️  Brightness control disabled. Install with: pip install screen-brightness-control")
    BRIGHTNESS_ENABLED = False

# ══════════════════════════════════════════════════
#  CONFIGURATION  —  keys loaded from config.py
# ══════════════════════════════════════════════════
ASSISTANT_NAME = "Nova"
WAKE_WORDS  = ["wake up nova", "wakeup nova", "wake nova"]
SLEEP_WORDS = ["go to sleep", "sleep mode", "sleep nova"]

# Load API keys from config.py (never uploaded to GitHub)
try:
    from config import GROQ_API_KEY, WEATHER_API_KEY, NEWS_API_KEY, WEATHER_CITY
except ImportError:
    print("⚠️  config.py not found!")
    print("    Copy config.example.py → config.py and fill in your API keys.")
    GROQ_API_KEY    = ""
    WEATHER_API_KEY = ""
    NEWS_API_KEY    = ""
    WEATHER_CITY    = "Gwalior"

# PyAutoGUI settings
pyautogui.PAUSE    = 0.3
pyautogui.FAILSAFE = True

# ── Text-to-Speech Setup ──────────────────────────
engine = pyttsx3.init()
voices = engine.getProperty('voices')
engine.setProperty('voice', voices[1].id if len(voices) > 1 else voices[0].id)
engine.setProperty('rate', 165)
engine.setProperty('volume', 1.0)

speech_lock = threading.Lock()

def speak(text: str, prefix=True):
    """Natural speaking with personality — thread-safe."""
    global engine
    print(f"\n🌟 Nova: {text}" if prefix else f"\n{text}")
    with speech_lock:
        try:
            engine.say(text)
            engine.runAndWait()
        except RuntimeError:
            engine = pyttsx3.init()
            voices = engine.getProperty('voices')
            engine.setProperty('voice', voices[1].id if len(voices) > 1 else voices[0].id)
            engine.setProperty('rate', 165)
            engine.setProperty('volume', 1.0)
            engine.say(text)
            engine.runAndWait()

# ── Speech Recognition Setup ─────────────────────
recognizer = sr.Recognizer()
recognizer.energy_threshold       = 300
recognizer.pause_threshold        = 0.8
recognizer.dynamic_energy_threshold = True

def listen(timeout=5) -> str:
    """Listen for voice input."""
    with sr.Microphone() as source:
        print("\n🎤 Listening...")
        recognizer.adjust_for_ambient_noise(source, duration=0.5)
        try:
            audio = recognizer.listen(source, timeout=timeout, phrase_time_limit=10)
        except sr.WaitTimeoutError:
            return ""
    try:
        text = recognizer.recognize_google(audio)
        print(f"👤 You: {text}")
        return text.lower()
    except sr.UnknownValueError:
        return ""
    except sr.RequestError:
        speak("Hmm, I'm having trouble with my speech service. Check your internet?")
        return ""

# ══════════════════════════════════════════════════
#  AI INTEGRATION — GROQ
# ══════════════════════════════════════════════════
ai_client = None

def initialize_ai():
    global ai_client
    if not AI_ENABLED:
        return False
    if GROQ_API_KEY == "YOUR_GROQ_API_KEY_HERE":
        print("\n⚠️  AI disabled: Please set your Groq API key")
        return False
    try:
        print("\n🔍 Connecting to Groq AI...")
        ai_client = Groq(api_key=GROQ_API_KEY)
        ai_client.chat.completions.create(
            messages=[{"role": "user", "content": "Say hi"}],
            model="llama-3.3-70b-versatile",
            max_tokens=10,
        )
        print("✅ AI powered by Groq (Llama 3.3 — Free & Fast!)")
        return True
    except Exception as e:
        print(f"⚠️  Groq AI initialization failed: {e}")
        return False

def ask_ai(question: str) -> str:
    if not ai_client:
        return None
    try:
        response = ai_client.chat.completions.create(
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are Nova, a friendly and helpful AI voice assistant "
                        "created by Anjali Saraswat. "
                        "Keep ALL responses concise — maximum 2 to 3 sentences. "
                        "Be conversational, warm, and helpful. "
                        "Never use bullet points or markdown formatting since your "
                        "response will be spoken aloud."
                    )
                },
                {"role": "user", "content": question}
            ],
            model="llama-3.3-70b-versatile",
            max_tokens=150,
            temperature=0.7,
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"AI Error: {e}")
        return None

# ══════════════════════════════════════════════════
#  HUMAN-LIKE RESPONSES
# ══════════════════════════════════════════════════
GREETINGS       = ["Hey there! What can I do for you?", "Hello! I'm all ears!",
                   "Hi! How can I help?", "Hey! What's up?"]
ACKNOWLEDGMENTS = ["Got it!", "On it!", "Sure thing!", "You got it!", "No problem!",
                   "Consider it done!", "Right away!", "Absolutely!"]
CONFIRMATIONS   = ["Done!", "All set!", "There you go!", "Finished!", "Complete!"]
ERRORS          = ["Oops, something went wrong there.",
                   "Hmm, that didn't work as expected.",
                   "Sorry, I hit a snag with that."]

# ══════════════════════════════════════════════════
#  TAB NAVIGATION  (FIXED)
#
#  FIX 1 — removed "close" from escape keywords
#           it was pressing Escape instead of closing apps
#  FIX 2 — removed "open" from enter/click keywords
#           it was pressing Enter instead of opening websites
#  FIX 3 — silence counter auto-exits after 3 timeouts
#           so it never freezes the UI thread indefinitely
# ══════════════════════════════════════════════════
def run_tab_navigation():
    """
    Dedicated voice loop for keyboard navigation.
    Works on browsers, desktop, any app window.
    Say: next / back / enter / click / space / escape / up / down / stop
    """
    speak("Tab navigation on! Say next, back, enter, space, escape, or stop to exit.")
    time.sleep(0.3)

    silence_count = 0   # FIX 3: track consecutive silences

    while True:
        command = listen(timeout=8)

        # FIX 3: auto-exit on repeated silence so the thread is never stuck forever
        if not command:
            silence_count += 1
            if silence_count >= 3:
                speak("Tab navigation timed out. Returning to normal mode.")
                break
            continue
        silence_count = 0

        numbers = re.findall(r'\d+', command)

        # ── Stop ──────────────────────────────────────
        if any(kw in command for kw in ["stop", "exit", "quit", "done",
                                         "end navigation", "tab mode off"]):
            speak("Navigation mode off!")
            break

        # ── Next (Tab forward) ────────────────────────
        elif "next" in command:
            times = int(numbers[0]) if numbers else command.count("next")
            times = max(1, times)
            for _ in range(times):
                pyautogui.press('tab')
                time.sleep(0.15)
            print(f"   ➡ Tab x{times}")

        # ── Back (Shift+Tab) ──────────────────────────
        elif any(kw in command for kw in ["back", "previous", "go back"]):
            times = int(numbers[0]) if numbers else 1
            for _ in range(times):
                pyautogui.hotkey('shift', 'tab')
                time.sleep(0.15)
            print(f"   ⬅ Shift+Tab x{times}")

        # ── Enter / Click ─────────────────────────────
        # FIX 2: removed "open" — was intercepting "open youtube" etc.
        elif any(kw in command for kw in ["enter", "click", "select",
                                           "ok", "press enter"]):
            pyautogui.press('enter')
            speak("Clicked!")

        # ── Space ─────────────────────────────────────
        elif "space" in command:
            pyautogui.press('space')
            print("   ⎵ Space")

        # ── Escape ────────────────────────────────────
        # FIX 1: removed "close" — was pressing Escape instead of closing apps
        elif any(kw in command for kw in ["escape", "cancel"]):
            pyautogui.press('escape')
            speak("Escaped!")

        # ── Arrow keys ────────────────────────────────
        elif "up" in command:
            times = int(numbers[0]) if numbers else 1
            pyautogui.press('up', presses=times)
            print(f"   ↑ Up x{times}")

        elif "down" in command:
            times = int(numbers[0]) if numbers else 1
            pyautogui.press('down', presses=times)
            print(f"   ↓ Down x{times}")

        elif "left" in command:
            pyautogui.press('left')
            print("   ← Left")

        elif "right" in command:
            pyautogui.press('right')
            print("   → Right")

        # ── Switch window ─────────────────────────────
        elif any(kw in command for kw in ["alt tab", "switch window", "switch app"]):
            pyautogui.hotkey('alt', 'tab')
            speak("Switching window!")

        # ── Fast tab ──────────────────────────────────
        elif any(kw in command for kw in ["fast tab", "hold tab", "quick tab"]):
            speak("Fast tabbing!")
            for _ in range(10):
                pyautogui.press('tab')
                time.sleep(0.08)

        else:
            speak("Say next, back, enter, or stop.")


def handle_tab_navigation(command: str) -> bool:
    """Trigger tab navigation mode from the main command processor."""
    if any(kw in command for kw in ["start tab mode", "tab mode", "navigate page",
                                     "start navigation", "navigation mode",
                                     "start tab navigation"]):
        run_tab_navigation()
        return True
    return False

# ══════════════════════════════════════════════════
#  WEATHER
# ══════════════════════════════════════════════════
def get_weather(command: str):
    """Fetch current weather. Supports 'weather in [city]'."""
    if not REQUESTS_ENABLED:
        speak("Weather needs the requests library. Run: pip install requests")
        return

    city = WEATHER_CITY
    for kw in ["weather in", "weather at", "weather for"]:
        if kw in command:
            city = command.split(kw, 1)[1].strip()
            break

    try:
        url = (f"http://api.openweathermap.org/data/2.5/weather"
               f"?q={city}&appid={WEATHER_API_KEY}&units=metric")
        response = requests.get(url, timeout=5)
        data     = response.json()

        if data.get("cod") != 200:
            speak(f"Sorry, I couldn't find weather data for {city}.")
            return

        temp      = round(data["main"]["temp"])
        feels     = round(data["main"]["feels_like"])
        humidity  = data["main"]["humidity"]
        desc      = data["weather"][0]["description"]
        city_name = data["name"]

        speak(
            f"Current weather in {city_name}: {desc}. "
            f"Temperature is {temp} degrees Celsius, "
            f"feels like {feels} degrees, "
            f"and humidity is {humidity} percent."
        )
    except Exception as e:
        speak("Sorry, I couldn't fetch the weather right now. Check your internet.")
        print(f"Weather error: {e}")

# ══════════════════════════════════════════════════
#  NEWS  (FIXED — function was completely missing before)
#
#  Fetches top 5 headlines from NewsAPI (India, English).
#  Trigger words: "news", "headlines", "what's happening"
# ══════════════════════════════════════════════════
def get_news(command: str = ""):
    """Fetch and read top news headlines using NewsAPI."""
    if not REQUESTS_ENABLED:
        speak("News needs the requests library. Run: pip install requests")
        return

    try:
        url = (f"https://newsapi.org/v2/top-headlines"
               f"?country=in&language=en&apiKey={NEWS_API_KEY}&pageSize=5")
        response = requests.get(url, timeout=6)
        data     = response.json()

        if data.get("status") != "ok":
            speak("Sorry, the news service returned an error. Please check your News API key.")
            return

        articles = [a for a in data.get("articles", []) if a.get("title")]
        if not articles:
            speak("I couldn't find any headlines right now. Try again in a moment.")
            return

        speak(f"Here are the top {len(articles)} headlines for you.")
        for i, article in enumerate(articles, 1):
            # Strip the " - Source Name" suffix that NewsAPI appends
            title = re.split(r'\s+-\s+', article["title"])[0].strip()
            if title:
                speak(f"Headline {i}: {title}")
                time.sleep(0.3)

    except Exception as e:
        speak("Sorry, couldn't fetch the news right now. Check your internet connection.")
        print(f"News error: {e}")

# ══════════════════════════════════════════════════
#  WHATSAPP SENDER
# ══════════════════════════════════════════════════
def send_whatsapp(command: str):
    """
    Send a WhatsApp message by contact name via the WhatsApp Desktop app.
    Say: 'send message to Papa I will be late'
    """
    contact_name = ""
    message      = ""

    for trigger in ["send message to", "send whatsapp to", "whatsapp to", "message to"]:
        if trigger in command:
            rest  = command.split(trigger, 1)[1].strip()
            parts = rest.split(" ", 1)
            contact_name = parts[0].strip()
            if len(parts) > 1:
                message = parts[1].strip()
            break

    if not contact_name:
        speak("Who should I send the message to?")
        contact_name = listen(timeout=6)
        if not contact_name:
            speak("I didn't catch the contact name.")
            return

    if not message:
        speak(f"What message should I send to {contact_name}?")
        message = listen(timeout=10)
        if not message:
            speak("I didn't catch the message.")
            return

    speak(f"Opening WhatsApp and sending message to {contact_name}.")

    whatsapp_paths = [
        os.path.join(os.environ.get("LOCALAPPDATA", ""), "WhatsApp", "WhatsApp.exe"),
        r"C:\Users\HP\AppData\Local\WhatsApp\WhatsApp.exe",
    ]
    opened = False
    for path in whatsapp_paths:
        if os.path.exists(path):
            subprocess.Popen(path)
            opened = True
            break

    if not opened:
        try:
            subprocess.Popen("WhatsApp.exe")
            opened = True
        except Exception:
            pass

    if not opened:
        subprocess.Popen(["cmd", "/c", "start", "whatsapp:"])

    time.sleep(4)
    pyautogui.hotkey('ctrl', 'f')
    time.sleep(1.5)
    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.3)
    pyautogui.write(contact_name, interval=0.08)
    time.sleep(2)
    pyautogui.press('down')
    time.sleep(0.5)
    pyautogui.press('enter')
    time.sleep(1.5)
    pyautogui.write(message, interval=0.05)
    time.sleep(0.5)
    pyautogui.press('enter')
    speak(f"Message sent to {contact_name}!")

# ══════════════════════════════════════════════════
#  MOOD-BASED MUSIC
# ══════════════════════════════════════════════════
MOOD_QUERIES = {
    "happy":    "happy upbeat music playlist",
    "sad":      "sad emotional music playlist",
    "chill":    "chill relaxing music playlist",
    "focus":    "deep focus study music",
    "workout":  "workout gym motivation music",
    "romantic": "romantic love songs playlist",
    "angry":    "aggressive metal rock music",
    "sleep":    "sleep calm relaxing music",
    "party":    "party dance hits playlist",
    "morning":  "good morning fresh start music",
}

def play_mood_music(command: str):
    """Play music based on mood by searching YouTube."""
    detected_mood = None
    for mood in MOOD_QUERIES:
        if mood in command:
            detected_mood = mood
            break

    if not detected_mood:
        speak("What mood are you in? Say happy, sad, chill, focus, workout, "
              "romantic, sleep, or party.")
        mood_response = listen()
        for mood in MOOD_QUERIES:
            if mood in mood_response:
                detected_mood = mood
                break

    if not detected_mood:
        speak("I didn't catch the mood. Try: play happy music or play chill music.")
        return

    query = MOOD_QUERIES[detected_mood]
    speak(f"Playing {detected_mood} music for you!")
    webbrowser.open(f"https://www.youtube.com/results?search_query={query.replace(' ', '+')}")

# ══════════════════════════════════════════════════
#  MATH CALCULATOR  (COMPLETELY REWRITTEN)
#
#  FIX 1 — "what is" alone no longer routes here.
#           Only routes when a math keyword is also present
#           (plus, minus, times, divided, percent, sqrt, power…)
#           so "what is the time" still goes to tell_time().
#
#  FIX 2 — Power operator (**) now works correctly.
#           Old code used re.sub that stripped one star from **,
#           turning "5 ** 2" into "5  2" which crashed eval().
#           New code builds the expression cleanly without regex stripping.
#
#  FIX 3 — Much broader word-to-symbol conversion so natural speech
#           like "five plus three" or "what is 30 percent of 500"
#           all work correctly.
#
#  FIX 4 — eval() is now sandboxed with only safe math functions
#           so no arbitrary Python can be executed.
# ══════════════════════════════════════════════════

# Words that confirm this is truly a math request
MATH_WORDS = [
    "plus", "add", "minus", "subtract", "times", "multiplied",
    "divided", "divide", "percent", "percentage", "square root",
    "sqrt", "power", "squared", "cubed", "modulo", "mod",
    "calculate", "compute", "solve",
]

def _words_to_number(word: str) -> str:
    """Convert English number words to digits."""
    ones  = {"zero":"0","one":"1","two":"2","three":"3","four":"4","five":"5",
             "six":"6","seven":"7","eight":"8","nine":"9","ten":"10",
             "eleven":"11","twelve":"12","thirteen":"13","fourteen":"14",
             "fifteen":"15","sixteen":"16","seventeen":"17","eighteen":"18",
             "nineteen":"19","twenty":"20","thirty":"30","forty":"40",
             "fifty":"50","sixty":"60","seventy":"70","eighty":"80","ninety":"90",
             "hundred":"100","thousand":"1000","million":"1000000"}
    return ones.get(word.lower(), word)

def _spoken_to_expr(text: str) -> str:
    """
    Convert a spoken math sentence into a Python expression string.
    e.g. "five plus three times two" → "5 + 3 * 2"
    """
    # Strip leading trigger words
    for kw in ["calculate", "what is", "compute", "solve", "find", "math"]:
        text = text.replace(kw, "").strip()

    # Convert number words to digits (word by word)
    tokens = text.split()
    tokens = [_words_to_number(t) for t in tokens]
    text   = " ".join(tokens)

    # Ordered replacements — longest phrases first to avoid partial matches
    replacements = [
        ("to the power of",  "**"),
        ("raised to",        "**"),
        ("multiplied by",    "*"),
        ("divided by",       "/"),
        ("divide by",        "/"),
        ("square root of",   "sqrt"),
        ("percent of",       "PERCENTOF"),
        ("percent",          "PERCENT"),
        ("plus",             "+"),
        ("add",              "+"),
        ("added to",         "+"),
        ("minus",            "-"),
        ("subtract",         "-"),
        ("subtracted from",  "-"),
        ("times",            "*"),
        ("power",            "**"),
        ("squared",          "**2"),
        ("cubed",            "**3"),
        ("modulo",           "%"),
        (" mod ",            "%"),
    ]
    for phrase, symbol in replacements:
        text = text.replace(phrase, symbol)

    return text.strip()

def calculate(command: str):
    """
    Solve any math problem from a spoken command.

    Supported:
      Basic:      "calculate 25 times 4"
      Percent:    "what is 15 percent of 200"
      Square root:"square root of 144"
      Power:      "2 to the power of 10"
      Number words:"five plus three"
      Complex:    "what is 100 divided by 4 plus 3 times 2"
    """

    # ── Percentage ────────────────────────────────
    # Pattern: "X percent of Y"
    pct_match = re.search(
        r'(\d+\.?\d*)\s*(?:percent(?:age)?|%)\s*of\s*(\d+\.?\d*)', command
    )
    if pct_match:
        pct    = float(pct_match.group(1))
        total  = float(pct_match.group(2))
        result = (pct / 100) * total
        result = int(result) if result == int(result) else round(result, 4)
        speak(f"{pct} percent of {total} is {result}.")
        return

    # ── Square root ───────────────────────────────
    sqrt_match = re.search(r'(?:square root of|sqrt)\s*(\d+\.?\d*)', command)
    if sqrt_match:
        num    = float(sqrt_match.group(1))
        result = math.sqrt(num)
        result = int(result) if result == int(result) else round(result, 6)
        speak(f"The square root of {num} is {result}.")
        return

    # ── General expression ────────────────────────
    expr = _spoken_to_expr(command)

    # Handle "PERCENTOF" placeholder that remains after conversion
    pct_expr = re.search(r'(\d+\.?\d*)\s*PERCENTOF\s*(\d+\.?\d*)', expr)
    if pct_expr:
        pct    = float(pct_expr.group(1))
        total  = float(pct_expr.group(2))
        result = (pct / 100) * total
        result = int(result) if result == int(result) else round(result, 4)
        speak(f"{pct} percent of {total} is {result}.")
        return

    # Handle lone "PERCENT" (e.g. "what is 20 percent" without "of")
    expr = re.sub(r'(\d+\.?\d*)\s*PERCENT', r'(\1/100)', expr)

    # Check there is at least one digit
    if not re.search(r'\d', expr):
        speak("Sorry, I couldn't find any numbers in that problem. Please try again.")
        return

    # Safe eval — only allow digits, operators, parens, dot, whitespace
    # Power (**) is preserved correctly here because we replace the word
    # before reaching this point, not inside a character class
    try:
        # Allowed characters: digits 0-9, operators + - * / % ** . ( ) space
        safe_expr = re.sub(r'[^0-9+\-*/.()% ]', '', expr).strip()

        if not safe_expr:
            speak("Sorry, I couldn't understand that expression.")
            return

        # Sandboxed eval: only math functions, no builtins
        safe_globals = {
            "__builtins__": {},
            "sqrt":  math.sqrt,
            "abs":   abs,
            "round": round,
            "pow":   pow,
            "pi":    math.pi,
            "e":     math.e,
            "sin":   math.sin,
            "cos":   math.cos,
            "tan":   math.tan,
            "log":   math.log,
            "log10": math.log10,
            "ceil":  math.ceil,
            "floor": math.floor,
        }
        result = eval(safe_expr, safe_globals, {})

        # Clean up output
        if isinstance(result, float):
            if result == int(result):
                result = int(result)
            else:
                result = round(result, 6)

        speak(f"The answer is {result}.")

    except ZeroDivisionError:
        speak("Oops! You can't divide by zero.")
    except SyntaxError:
        speak("I couldn't parse that expression. Try saying it differently — "
              "for example: calculate 25 times 4, or 15 percent of 200.")
    except Exception as ex:
        speak("Sorry, I couldn't solve that. Try something like: "
              "calculate 25 times 4, or square root of 144.")
        print(f"Calc error on '{expr}': {ex}")

def _is_math_command(command: str) -> bool:
    """
    Returns True only when the command is genuinely a math request.
    Prevents 'what is the time' from routing to the calculator.
    """
    # Direct trigger words always mean math
    for kw in ["calculate", "compute", "solve", "square root", "sqrt"]:
        if kw in command:
            return True

    # "percent of" pattern
    if re.search(r'\d+\s*(?:percent|%)\s*of', command):
        return True

    # "what is" only counts as math if a math operator word is also present
    if "what is" in command:
        math_operators = [
            "plus", "minus", "times", "divided", "multiplied",
            "percent", "power", "squared", "cubed", "sqrt",
            "square root", "add", "subtract", "multiply",
        ]
        return any(op in command for op in math_operators)

    return False

# ══════════════════════════════════════════════════
#  LANGUAGE TRANSLATOR
# ══════════════════════════════════════════════════
LANGUAGE_CODES = {
    "hindi": "hi",      "spanish": "es",    "french": "fr",
    "german": "de",     "japanese": "ja",   "chinese": "zh-cn",
    "arabic": "ar",     "russian": "ru",    "portuguese": "pt",
    "italian": "it",    "korean": "ko",     "urdu": "ur",
    "bengali": "bn",    "tamil": "ta",      "telugu": "te",
    "marathi": "mr",    "gujarati": "gu",   "punjabi": "pa",
    "english": "en",
}

def translate_text(command: str):
    """Translate spoken text to another language."""
    if not TRANSLATOR_ENABLED:
        speak("Translator needs deep-translator. Run: pip install deep-translator")
        return

    target_lang = None
    target_name = None
    for lang_name, lang_code in LANGUAGE_CODES.items():
        if lang_name in command:
            target_lang = lang_code
            target_name = lang_name
            break

    if not target_lang:
        speak("Which language? Say Hindi, French, Spanish, German, Japanese, etc.")
        lang_response = listen()
        for lang_name, lang_code in LANGUAGE_CODES.items():
            if lang_name in lang_response:
                target_lang = lang_code
                target_name = lang_name
                break

    if not target_lang:
        speak("I couldn't detect the language. Please try again.")
        return

    for kw in ["translate", "say", "in " + (target_name or ""),
               "to " + (target_name or "")]:
        command = command.replace(kw, "")
    text_to_translate = command.strip()

    if not text_to_translate or len(text_to_translate) < 2:
        speak("What would you like me to translate?")
        text_to_translate = listen(timeout=8)
        if not text_to_translate:
            return

    try:
        translated = GoogleTranslator(source='auto', target=target_lang).translate(text_to_translate)
        speak(f"In {target_name}, that is: {translated}")
    except Exception as e:
        speak(f"{random.choice(ERRORS)} Translation failed. Try again.")
        print(f"Translation error: {e}")

# ══════════════════════════════════════════════════
#  APP CLOSER
# ══════════════════════════════════════════════════
APP_PROCESS_MAP = {
    "chrome":        ["chrome.exe"],
    "firefox":       ["firefox.exe"],
    "edge":          ["msedge.exe"],
    "spotify":       ["spotify.exe"],
    "notepad":       ["notepad.exe"],
    "calculator":    ["calculatorapp.exe", "calc.exe"],
    "word":          ["winword.exe"],
    "excel":         ["excel.exe"],
    "powerpoint":    ["powerpnt.exe"],
    "vlc":           ["vlc.exe"],
    "vs code":       ["code.exe"],
    "visual studio": ["code.exe"],
    "paint":         ["mspaint.exe"],
    "task manager":  ["taskmgr.exe"],
    "explorer":      ["explorer.exe"],
    "whatsapp":      ["whatsapp.exe"],
    "discord":       ["discord.exe"],
    "zoom":          ["zoom.exe"],
    "teams":         ["teams.exe"],
}

def close_application(command: str) -> bool:
    """Close a running application by name."""
    if not PSUTIL_ENABLED:
        speak("App closer needs psutil. Run: pip install psutil")
        return False

    app_to_close = None
    for app_name in APP_PROCESS_MAP:
        if app_name in command:
            app_to_close = app_name
            break

    if not app_to_close:
        return False

    closed = False
    for proc in psutil.process_iter(['name']):
        try:
            proc_name = proc.info['name'].lower()
            if proc_name in [p.lower() for p in APP_PROCESS_MAP[app_to_close]]:
                proc.terminate()
                closed = True
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass

    if closed:
        speak(f"Closed {app_to_close}!")
    else:
        speak(f"{app_to_close.capitalize()} doesn't seem to be running.")
    return True

# ══════════════════════════════════════════════════
#  SCREEN BRIGHTNESS
# ══════════════════════════════════════════════════
def control_brightness(command: str) -> bool:
    """Control screen brightness by voice."""
    if not BRIGHTNESS_ENABLED:
        speak("Brightness needs screen-brightness-control. "
              "Run: pip install screen-brightness-control")
        return False

    try:
        set_match = re.search(r'set brightness to (\d+)', command)
        if set_match:
            level = max(10, min(100, int(set_match.group(1))))
            sbc.set_brightness(level)
            speak(f"Brightness set to {level} percent.")
            return True

        if "increase brightness" in command or "brightness up" in command:
            new_level = min(100, sbc.get_brightness()[0] + 20)
            sbc.set_brightness(new_level)
            speak(f"Brightness increased to {new_level} percent.")
            return True

        if any(kw in command for kw in ["decrease brightness", "brightness down",
                                         "reduce brightness"]):
            new_level = max(10, sbc.get_brightness()[0] - 20)
            sbc.set_brightness(new_level)
            speak(f"Brightness decreased to {new_level} percent.")
            return True

        if any(kw in command for kw in ["max brightness", "full brightness"]):
            sbc.set_brightness(100)
            speak("Brightness set to maximum!")
            return True

        if any(kw in command for kw in ["min brightness", "low brightness",
                                         "dim screen"]):
            sbc.set_brightness(20)
            speak("Brightness set to low.")
            return True

        if ("what is" in command and "brightness" in command) \
                or "check brightness" in command:
            current = sbc.get_brightness()[0]
            speak(f"Current brightness is {current} percent.")
            return True

    except Exception as e:
        speak("Sorry, I couldn't control the brightness on this display.")
        print(f"Brightness error: {e}")
        return True

    return False

# ══════════════════════════════════════════════════
#  APP AUTOMATION
# ══════════════════════════════════════════════════
def type_text(text: str):
    time.sleep(0.5)
    pyautogui.write(text, interval=0.05)

def press_key(key: str, times: int = 1):
    for _ in range(times):
        pyautogui.press(key)
        time.sleep(0.1)

def youtube_automation(command: str):
    if "play" in command:
        query  = command.split("play", 1)[1].strip()
        action = "play"
    elif "search for" in command or "search" in command:
        query  = command.replace("search for", "").replace("search", "").strip()
        action = "search"
    else:
        speak(f"{random.choice(ACKNOWLEDGMENTS)} Opening YouTube.")
        webbrowser.open("https://www.youtube.com")
        return

    if not query:
        speak("What would you like me to search for on YouTube?")
        query = listen()
        if not query:
            return

    speak(f"{random.choice(ACKNOWLEDGMENTS)} Opening YouTube and searching for {query}.")
    webbrowser.open("https://www.youtube.com")
    time.sleep(3)
    pyautogui.press('/')
    time.sleep(0.5)
    type_text(query)
    time.sleep(0.5)
    pyautogui.press('enter')

    if action == "play":
        time.sleep(2)
        speak("Playing the first result!")
        pyautogui.press('tab', presses=46)
        pyautogui.press('enter')

def spotify_control(command: str):
    if "open spotify" in command:
        query = ""
        if "play" in command:
            query = command.split("play", 1)[1].strip()
        speak(f"{random.choice(ACKNOWLEDGMENTS)} Opening Spotify.")
        try:
            subprocess.Popen("spotify.exe")
            time.sleep(2)
        except Exception:
            webbrowser.open("https://open.spotify.com")
            time.sleep(3)
        if query:
            speak(f"Searching for {query}.")
            time.sleep(1)
            pyautogui.hotkey('ctrl', 'l')
            time.sleep(0.5)
            type_text(query)
            time.sleep(0.5)
            pyautogui.press('enter')
            time.sleep(1)
            pyautogui.press('enter')
    elif "play" in command or "resume" in command:
        speak("Playing!"); pyautogui.press('playpause')
    elif "pause" in command or "stop" in command:
        speak("Pausing!"); pyautogui.press('playpause')
    elif "next" in command:
        speak("Next track!"); pyautogui.press('nexttrack')
    elif "previous" in command or "last" in command:
        speak("Previous track!"); pyautogui.press('prevtrack')

def media_control(command: str) -> bool:
    if "play" in command or "resume" in command:
        speak("Playing!"); pyautogui.press('playpause'); return True
    elif "pause" in command:
        speak("Pausing!"); pyautogui.press('playpause'); return True
    elif "next song" in command or "next track" in command or "skip" in command:
        speak("Next track!"); pyautogui.press('nexttrack'); return True
    elif "previous song" in command or "previous track" in command:
        speak("Previous track!"); pyautogui.press('prevtrack'); return True
    elif "stop music" in command or "stop playing" in command:
        speak("Stopping!"); pyautogui.press('stop'); return True
    elif "full screen" in command or "fullscreen" in command:
        speak("Going full screen!"); pyautogui.press('f'); return True
    elif "exit full screen" in command:
        speak("Exiting full screen!"); pyautogui.press('escape'); return True
    return False

def browser_control(command: str) -> bool:
    if "new tab" in command or "open new tab" in command:
        speak("Opening new tab!"); pyautogui.hotkey('ctrl', 't'); return True
    elif "close tab" in command:
        speak("Closing tab!"); pyautogui.hotkey('ctrl', 'w'); return True
    elif "refresh" in command or "reload" in command:
        speak("Refreshing page!"); pyautogui.press('f5'); return True
    elif "go back" in command and "browser" in command:
        speak("Going back!"); pyautogui.hotkey('alt', 'left'); return True
    elif "go forward" in command:
        speak("Going forward!"); pyautogui.hotkey('alt', 'right'); return True
    elif "zoom in" in command:
        speak("Zooming in!"); pyautogui.hotkey('ctrl', '+'); return True
    elif "zoom out" in command:
        speak("Zooming out!"); pyautogui.hotkey('ctrl', '-'); return True
    return False

def type_command(command: str):
    text_to_type = command.replace("type", "").strip()
    if not text_to_type:
        speak("What would you like me to type?")
        text_to_type = listen()
        if not text_to_type:
            return
    speak(f"Typing: {text_to_type}")
    time.sleep(0.5)
    type_text(text_to_type)

# ══════════════════════════════════════════════════
#  TIME & DATE
# ══════════════════════════════════════════════════
def tell_time():
    now    = datetime.datetime.now()
    hour   = now.strftime('%I').lstrip('0')
    minute = now.strftime('%M')
    period = now.strftime('%p')
    speak(random.choice([
        f"It's {hour} {minute} {period}.",
        f"The time is {hour} {minute} {period}.",
        f"Right now it's {hour} {minute} {period}.",
    ]))

def tell_date():
    today = datetime.datetime.now()
    day   = today.strftime('%A')
    month = today.strftime('%B')
    date  = today.strftime('%d').lstrip('0')
    year  = today.strftime('%Y')
    speak(random.choice([
        f"Today is {day}, {month} {date}, {year}.",
        f"It's {day}, {month} {date}, {year}.",
    ]))

# ══════════════════════════════════════════════════
#  OPEN APPLICATIONS
# ══════════════════════════════════════════════════
def find_office_app(app_name):
    office_paths = [
        r"C:\Program Files\Microsoft Office\root\Office16",
        r"C:\Program Files (x86)\Microsoft Office\root\Office16",
        r"C:\Program Files\Microsoft Office\Office16",
        r"C:\Program Files (x86)\Microsoft Office\Office16",
        r"C:\Program Files\Microsoft Office 15\root\Office15",
        r"C:\Program Files (x86)\Microsoft Office 15\root\Office15",
    ]
    app_exes = {"word": "WINWORD.EXE", "excel": "EXCEL.EXE",
                "powerpoint": "POWERPNT.EXE"}
    if app_name not in app_exes:
        return None
    for base_path in office_paths:
        full_path = os.path.join(base_path, app_exes[app_name])
        if os.path.exists(full_path):
            return full_path
    return None

APP_MAP = {
    "notepad": "notepad.exe", "calculator": "calc.exe", "paint": "mspaint.exe",
    "file explorer": "explorer.exe", "task manager": "taskmgr.exe",
    "command prompt": "cmd.exe", "settings": "ms-settings:",
    "camera": "microsoft.windows.camera:",
    "chrome": "chrome.exe", "firefox": "firefox.exe", "vlc": "vlc.exe",
    "edge": "msedge.exe", "spotify": "spotify.exe",
    "vs code": "code.exe", "visual studio code": "code.exe",
}

def open_application(command: str) -> bool:
    office_apps = {
        "word": "word", "excel": "excel",
        "powerpoint": "powerpoint", "power point": "powerpoint"
    }
    for app_name, office_key in office_apps.items():
        if app_name in command:
            speak(f"{random.choice(ACKNOWLEDGMENTS)} Opening {app_name}.")
            app_path = find_office_app(office_key)
            if app_path:
                try:
                    subprocess.Popen([app_path])
                    speak(random.choice(CONFIRMATIONS))
                    return True
                except Exception as e:
                    print(f"Error: {e}")
            try:
                os.startfile({"word":"winword","excel":"excel",
                              "powerpoint":"powerpnt"}[office_key])
                speak(random.choice(CONFIRMATIONS))
                return True
            except Exception:
                speak(f"{random.choice(ERRORS)} Couldn't open {app_name}.")
            return True

    for app_name, app_cmd in APP_MAP.items():
        if app_name in command:
            speak(f"{random.choice(ACKNOWLEDGMENTS)} Opening {app_name}.")
            try:
                if app_cmd.endswith(":"): os.startfile(app_cmd)
                else:                     subprocess.Popen(app_cmd)
                speak(random.choice(CONFIRMATIONS))
            except Exception:
                speak(f"{random.choice(ERRORS)} Couldn't open {app_name}.")
            return True
    return False

# ══════════════════════════════════════════════════
#  FILE SEARCH
# ══════════════════════════════════════════════════
SEARCH_LOCATIONS = [
    os.path.expanduser("~\\Desktop"),   os.path.expanduser("~\\Documents"),
    os.path.expanduser("~\\Downloads"), os.path.expanduser("~\\Music"),
    os.path.expanduser("~\\Pictures"),  os.path.expanduser("~\\Videos"),
]

def search_and_open_file(command: str):
    filename_query = ""
    for trigger in ["search file", "find file", "open file", "find my file"]:
        if trigger in command:
            filename_query = command.replace(trigger, "").strip()
            break
    if not filename_query:
        speak("What file are you looking for?")
        filename_query = listen()
        if not filename_query:
            return

    speak(f"Searching for {filename_query}...")
    found_files = []
    for folder in SEARCH_LOCATIONS:
        if not os.path.exists(folder): continue
        matches = glob.glob(
            os.path.join(folder, "**", f"*{filename_query}*"), recursive=True
        )
        found_files.extend(matches)

    if not found_files:
        speak(f"Sorry, I couldn't find any file with '{filename_query}'.")
        return

    speak(f"Found {len(found_files)} file{'s' if len(found_files)>1 else ''}. "
          f"Opening {os.path.basename(found_files[0])}.")
    try:
        os.startfile(found_files[0])
    except Exception:
        speak(f"{random.choice(ERRORS)} Found it but couldn't open it.")

def open_folder(command: str) -> bool:
    folders = {
        "desktop":   os.path.expanduser("~\\Desktop"),
        "documents": os.path.expanduser("~\\Documents"),
        "downloads": os.path.expanduser("~\\Downloads"),
        "pictures":  os.path.expanduser("~\\Pictures"),
        "music":     os.path.expanduser("~\\Music"),
        "videos":    os.path.expanduser("~\\Videos"),
    }
    for folder_name, folder_path in folders.items():
        if folder_name in command and "open" in command:
            speak(f"{random.choice(ACKNOWLEDGMENTS)} Opening {folder_name}.")
            subprocess.Popen(f'explorer "{folder_path}"')
            return True
    return False

# ══════════════════════════════════════════════════
#  WEB SEARCH
# ══════════════════════════════════════════════════
def search_web(command: str):
    sites = {
        "facebook":  "https://www.facebook.com",
        "instagram": "https://www.instagram.com",
        "twitter":   "https://www.twitter.com",
        "github":    "https://www.github.com",
        "gmail":     "https://mail.google.com",
        "maps":      "https://maps.google.com",
        "wikipedia": "https://www.wikipedia.org",
        "reddit":    "https://www.reddit.com",
        "amazon":    "https://www.amazon.com",
        "netflix":   "https://www.netflix.com",
    }
    for site_name, site_url in sites.items():
        if f"open {site_name}" in command:
            speak(f"{random.choice(ACKNOWLEDGMENTS)} Opening {site_name}.")
            webbrowser.open(site_url)
            return

    if "search" in command or "google" in command:
        query = (command.replace("search", "").replace("google", "")
                        .replace("for", "").strip())
        if query:
            speak(f"Searching for {query}.")
            webbrowser.open(f"https://www.google.com/search?q={query}")
        else:
            speak("What would you like me to search for?")
            query = listen()
            if query:
                speak("Looking that up now!")
                webbrowser.open(f"https://www.google.com/search?q={query}")

# ══════════════════════════════════════════════════
#  REMINDERS
# ══════════════════════════════════════════════════
def set_reminder(command: str):
    speak("What should I remind you about?")
    reminder_text = listen() or "your reminder"
    speak("In how many minutes?")
    minutes_text = listen()
    minutes = None
    for word in minutes_text.split():
        if word.isdigit():
            minutes = int(word)
            break
    if minutes is None:
        speak("I didn't catch the time. Please say a number like 5 or 10.")
        return
    speak(f"Alright! I'll remind you about {reminder_text} in {minutes} minutes.")
    def reminder_thread():
        time.sleep(minutes * 60)
        speak(f"Hey! Time's up! {reminder_text}")
    threading.Thread(target=reminder_thread, daemon=True).start()

# ══════════════════════════════════════════════════
#  EXCEL SPREADSHEET
# ══════════════════════════════════════════════════
def create_excel_spreadsheet(command: str):
    if not EXCEL_ENABLED:
        speak("Excel creation needs openpyxl. Run: pip install openpyxl")
        return

    words    = command.split()
    num_cols = 5
    num_rows = 10
    topic    = "Data"

    for i, word in enumerate(words):
        if word.isdigit():
            number = int(word)
            if i + 1 < len(words):
                next_word = words[i + 1]
                if "column" in next_word:
                    num_cols = number
                elif any(k in next_word for k in ["row","student","employee","item"]):
                    num_rows = number
                else:
                    num_cols = number
            else:
                num_cols = number

    if "student"  in command: topic = "Student"
    elif "employee" in command: topic = "Employee"
    elif "product"  in command: topic = "Product"
    elif "sales"    in command: topic = "Sales"

    speak(f"{random.choice(ACKNOWLEDGMENTS)} Creating an Excel spreadsheet "
          f"with {num_cols} columns and {num_rows} rows.")

    try:
        wb    = Workbook()
        sheet = wb.active
        sheet.title = f"{topic} Data"

        header_templates = {
            2: ["Name","Value"],
            3: ["Name","Category","Value"],
            4: ["S.No","Name","Category","Value"],
            5: ["S.No","Name","Age","Category","Marks"],
            6: ["S.No","Name","Age","Gender","Category","Marks"],
            7: ["S.No","Name","Age","Gender","Department","Category","Marks"],
            8: ["S.No","Name","Age","Gender","Department","Category","Marks","Grade"],
        }
        headers = header_templates.get(num_cols,
            ["S.No"] + [f"Column {i}" for i in range(1, num_cols)])

        if topic == "Student" and num_cols >= 5:
            headers = ["S.No","Name","Age","Class","Marks"]
            if num_cols > 5:
                headers.extend([f"Subject {i}" for i in range(1, num_cols-4)])
        elif topic == "Employee" and num_cols >= 5:
            headers = ["S.No","Name","Age","Department","Salary"]
            if num_cols > 5:
                headers.extend([f"Field {i}" for i in range(1, num_cols-4)])

        headers     = headers[:num_cols]
        hdr_font    = Font(bold=True, color="FFFFFF", size=12)
        hdr_fill    = PatternFill(start_color="0D9488", end_color="0D9488",
                                  fill_type="solid")
        hdr_align   = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value, cell.font = header, hdr_font
            cell.fill, cell.alignment, cell.border = hdr_fill, hdr_align, thin_border
            sheet.column_dimensions[cell.column_letter].width = 15

        for row_idx in range(2, num_rows + 2):
            for col_idx in range(1, num_cols + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                h    = headers[col_idx - 1]
                if col_idx == 1:               cell.value = row_idx - 1
                elif h == "Name":              cell.value = f"{topic} {row_idx-1}"
                elif h == "Age":               cell.value = 20 + (row_idx % 10)
                elif h in ["Marks","Value","Salary"]:
                    cell.value = 70 + (row_idx % 30)
                else:                          cell.value = ""
                cell.border    = thin_border
                cell.alignment = Alignment(horizontal="center")

        if any(h in ["Marks","Value","Salary"] for h in headers):
            total_row = num_rows + 2
            sheet.cell(row=total_row, column=1).value = "Total"
            sheet.cell(row=total_row, column=1).font  = Font(bold=True)
            for col_idx, header in enumerate(headers, start=1):
                if header in ["Marks","Value","Salary"]:
                    cell       = sheet.cell(row=total_row, column=col_idx)
                    cell.value = (f"=SUM({cell.column_letter}2:"
                                  f"{cell.column_letter}{num_rows+1})")
                    cell.font  = Font(bold=True)
                    cell.border = thin_border

        sheet.freeze_panes = "A2"
        filename = (f"{topic}_Spreadsheet_"
                    f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        filepath = os.path.join(os.path.expanduser("~\\Desktop"), filename)
        wb.save(filepath)
        speak(f"Spreadsheet created! Saved as {filename} on your Desktop.")
        try:
            os.startfile(filepath)
        except Exception:
            pass
    except Exception as e:
        speak(f"{random.choice(ERRORS)} Couldn't create the spreadsheet.")
        print(f"Excel error: {e}")

# ══════════════════════════════════════════════════
#  SYSTEM CONTROLS
# ══════════════════════════════════════════════════
def system_control(command: str) -> bool:
    if "scroll down" in command or "page down" in command:
        scroll_amount = next((int(w) for w in command.split() if w.isdigit()), 3)
        speak(f"{random.choice(ACKNOWLEDGMENTS)} Scrolling down.")
        for _ in range(scroll_amount):
            pyautogui.scroll(-300); time.sleep(0.1)
        return True

    elif "scroll up" in command or "page up" in command:
        scroll_amount = next((int(w) for w in command.split() if w.isdigit()), 3)
        speak(f"{random.choice(ACKNOWLEDGMENTS)} Scrolling up.")
        for _ in range(scroll_amount):
            pyautogui.scroll(300); time.sleep(0.1)
        return True

    elif "screenshot" in command:
        speak(f"{random.choice(ACKNOWLEDGMENTS)} Taking a screenshot.")
        try:
            screenshot = pyautogui.screenshot()
            filename   = f"screenshot_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            pictures   = os.path.expanduser("~\\Pictures")
            filepath   = os.path.join(
                pictures if os.path.exists(pictures)
                else os.path.expanduser("~\\Desktop"), filename
            )
            screenshot.save(filepath)
            speak("Screenshot saved!")
        except Exception:
            speak("Sorry, couldn't save the screenshot.")
        return True

    elif "volume up" in command or "increase volume" in command:
        speak("Turning it up!")
        [pyautogui.press('volumeup') for _ in range(5)]
        return True

    elif "volume down" in command or "decrease volume" in command:
        speak("Turning it down!")
        [pyautogui.press('volumedown') for _ in range(5)]
        return True

    elif "mute" in command or "unmute" in command:
        speak(f"{random.choice(ACKNOWLEDGMENTS)}")
        pyautogui.press('volumemute')
        return True

    elif "lock" in command and ("computer" in command or "pc" in command):
        speak("Locking your computer. See you soon!")
        subprocess.run(["rundll32.exe", "user32.dll,LockWorkStation"])
        return True

    elif "shutdown" in command or "shut down" in command:
        speak("Are you sure you want to shut down? Say yes to confirm.")
        if "yes" in listen():
            speak("Shutting down. Take care!"); os.system("shutdown /s /t 5")
        else:
            speak("Shutdown cancelled.")
        return True

    elif "restart" in command:
        speak("Are you sure you want to restart? Say yes to confirm.")
        if "yes" in listen():
            speak("Restarting now!"); os.system("shutdown /r /t 5")
        else:
            speak("Restart cancelled.")
        return True

    return False

# ══════════════════════════════════════════════════
#  CONVERSATIONS
# ══════════════════════════════════════════════════
def handle_conversation(command: str) -> bool:
    if any(w in command for w in ["hello", "hi", "hey"]):
        speak(random.choice(GREETINGS)); return True
    elif "how are you" in command or "how's it going" in command:
        speak(random.choice([
            "I'm doing great! How about you?",
            "Fantastic! What can I do for you?",
            "All systems go! What do you need?",
        ])); return True
    elif "your name" in command or "who are you" in command:
        speak(random.choice([
            "I'm Nova, your AI assistant! Created by Anjali Saraswat.",
            "My name is Nova! Here to make your life easier!",
        ])); return True
    elif "what can you do" in command or "your capabilities" in command \
            or "help" in command:
        speak(
            "I can open apps, search files, browse the web, control media, "
            "use tab navigation, check weather, read news, send WhatsApp messages, "
            "play mood music, solve math, translate languages, close apps, "
            "control brightness, and answer questions using AI. Just ask!"
        ); return True
    elif "thank" in command:
        speak(random.choice(["You're very welcome!","Happy to help!",
                              "Anytime!","My pleasure!"])); return True
    elif "good morning" in command:
        speak("Good morning! Hope you have an amazing day!"); return True
    elif "good afternoon" in command:
        speak("Good afternoon! How's your day going?"); return True
    elif "good night" in command or "goodnight" in command:
        speak("Good night! Sleep well!"); return True
    elif "joke" in command:
        jokes = [
            "Why do programmers prefer dark mode? Because light attracts bugs!",
            "Why was the computer cold? It left its Windows open!",
            "Why do Java developers wear glasses? Because they don't C sharp!",
            "Why did the developer go broke? Because he used up all his cache!",
            "What's a computer's favorite snack? Microchips!",
        ]
        speak(random.choice(jokes)); return True
    elif "are you real" in command or "are you human" in command:
        speak("I'm an AI, so not human, but I'm here for real to help you!"); return True
    return False

# ══════════════════════════════════════════════════
#  COMMAND PROCESSOR  (FIXED routing order)
#
#  KEY CHANGES:
#  1. Time & Date checks moved BEFORE calculator
#     so "what time is it" and "what is today" work correctly
#  2. Calculator now uses _is_math_command() which only returns
#     True when a genuine math operator word is present alongside
#     "what is" — preventing any time/date/general question from
#     being misrouted to the calculator
#  3. News added with dedicated keywords
# ══════════════════════════════════════════════════
def process_command(command: str):
    if not command:
        return

    # ── Exit ─────────────────────────────────────
    if any(w in command for w in ["exit", "quit", "bye", "goodbye", "close nova"]):
        speak("Goodbye! It was nice talking to you!")
        sys.exit(0)

    # ── Tab Navigation ────────────────────────────
    elif handle_tab_navigation(command):
        pass

    # ── Weather ───────────────────────────────────
    elif any(kw in command for kw in ["weather", "temperature", "forecast",
                                       "will it rain"]):
        get_weather(command)

    # ── News (FIXED — was missing entirely) ───────
    elif any(kw in command for kw in ["news", "headlines", "what's happening",
                                       "read the news", "read me the news",
                                       "latest news"]):
        get_news(command)

    # ── WhatsApp ──────────────────────────────────
    elif any(kw in command for kw in ["send message to", "send whatsapp",
                                       "whatsapp to", "message to"]):
        send_whatsapp(command)

    # ── Mood Music ────────────────────────────────
    elif any(kw in command for kw in ["play happy", "play sad", "play chill",
                                       "play focus", "play workout", "play romantic",
                                       "play sleep music", "play party",
                                       "play morning", "mood music", "play angry"]):
        play_mood_music(command)

    # ── Time (MOVED UP — before calculator) ───────
    # Must come before calculator because "what is the time" must NOT
    # go to calculate()
    elif "time" in command and not _is_math_command(command):
        tell_time()

    # ── Date (MOVED UP — before calculator) ───────
    elif any(kw in command for kw in ["date", "today", "what day"]) \
            and not _is_math_command(command):
        tell_date()

    # ── Calculator (FIXED) ────────────────────────
    # Uses _is_math_command() so "what is the time/date" never reaches here
    elif _is_math_command(command):
        calculate(command)

    # ── Translator ────────────────────────────────
    elif any(kw in command for kw in ["translate", "say in",
                                       "say it in", "how do you say"]):
        translate_text(command)

    # ── App Closer ────────────────────────────────
    elif any(kw in command for kw in ["close", "kill", "shut"]) \
            and any(app in command for app in APP_PROCESS_MAP):
        close_application(command)

    # ── Brightness ────────────────────────────────
    elif "brightness" in command:
        control_brightness(command)

    # ── Excel Spreadsheet ─────────────────────────
    elif any(kw in command for kw in ["create spreadsheet", "make spreadsheet",
                                       "create excel", "make excel"]):
        create_excel_spreadsheet(command)

    # ── YouTube ───────────────────────────────────
    elif "youtube" in command and any(kw in command for kw in ["play", "search"]):
        youtube_automation(command)

    # ── Spotify ───────────────────────────────────
    elif "spotify" in command or ("play" in command and "music" in command):
        spotify_control(command)

    # ── Media controls ────────────────────────────
    elif media_control(command):
        pass

    # ── Browser controls ──────────────────────────
    elif browser_control(command):
        pass

    # ── Type text ─────────────────────────────────
    elif command.startswith("type "):
        type_command(command)

    # ── Reminders ─────────────────────────────────
    elif "remind" in command or "reminder" in command:
        set_reminder(command)

    # ── System controls ───────────────────────────
    elif system_control(command):
        pass

    # ── File operations ───────────────────────────
    elif any(kw in command for kw in ["search file", "find file",
                                       "open file", "find my file"]):
        search_and_open_file(command)

    elif open_folder(command):
        pass

    # ── Open applications ─────────────────────────
    elif any(app in command for app in
             list(APP_MAP.keys()) + ["word", "excel", "powerpoint", "power point"]):
        open_application(command)

    # ── Web search ────────────────────────────────
    elif any(kw in command for kw in ["search", "google", "open facebook",
                                       "open instagram", "open github",
                                       "open gmail", "open maps", "open wikipedia",
                                       "open reddit", "open amazon", "open netflix"]):
        search_web(command)

    # ── Conversation ──────────────────────────────
    elif handle_conversation(command):
        pass

    # ── AI fallback ───────────────────────────────
    else:
        if ai_client:
            ai_response = ask_ai(command)
            if ai_response:
                speak(ai_response)
            else:
                speak("Hmm, I'm not sure about that. Could you rephrase?")
        else:
            speak("I'm not sure how to help with that. "
                  "Try saying 'help' to see what I can do!")

# ══════════════════════════════════════════════════
#  MAIN LOOPS
# ══════════════════════════════════════════════════
def run_nova():
    speak(f"Hello! I'm {ASSISTANT_NAME}, your AI assistant!")
    speak("Say 'wake up Nova' anytime to activate me!")
    is_awake = False

    while True:
        if not is_awake:
            print(f"\n💤 {ASSISTANT_NAME} is sleeping… (Say 'wake up nova' to activate)")
            trigger = listen(timeout=10)
            if any(wake in trigger for wake in WAKE_WORDS):
                is_awake = True
                speak(random.choice([
                    "Yes? I'm listening!", "I'm here! What can I do?",
                    "Hello! How can I help?", "I'm awake!",
                ]))
        else:
            command = listen()
            if not command:
                continue
            if any(sleep in command for sleep in SLEEP_WORDS):
                is_awake = False
                speak(random.choice([
                    "Going to sleep. Wake me when you need me!",
                    "Taking a nap. Just say my name!",
                ]))
            else:
                process_command(command)

def run_simple_mode():
    speak(f"Hello! I'm {ASSISTANT_NAME}! Always listening. Say 'exit' when done!")
    while True:
        command = listen()
        process_command(command)

# ══════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════
if __name__ == "__main__":
    print("=" * 65)
    print(f"   🌟 {ASSISTANT_NAME.upper()} — AI VOICE ASSISTANT v3.1 (FIXED) 🌟")
    print("   By: Anjali Saraswat")
    print("   AI: Groq — llama-3.3-70b-versatile (Free!)")
    print("=" * 65)

    print("\n  📦 INSTALL ALL DEPENDENCIES:")
    print("  pip install SpeechRecognition pyttsx3 pyautogui groq openpyxl")
    print("  pip install requests psutil screen-brightness-control deep-translator")

    print("\n  🆕 FEATURES:")
    print("  🖱️  Tab Navigation  → 'start tab mode', then say next/back/enter/stop")
    print("  🌤️  Weather         → 'what\\'s the weather' / 'weather in Mumbai'")
    print("  📰  News            → 'read me the news' / 'latest headlines'")
    print("  💬  WhatsApp        → 'send message to Papa I\\'ll be late'")
    print("  🎵  Mood Music      → 'play happy music' / 'play chill music'")
    print("  🧮  Calculator      → 'calculate 25 times 4' / '15 percent of 200'")
    print("  🌍  Translator      → 'translate hello to Hindi'")
    print("  ❌  App Closer      → 'close chrome' / 'close spotify'")
    print("  💡  Brightness      → 'increase brightness' / 'set brightness to 70'")

    if AI_ENABLED:
        ai_initialized = initialize_ai()
        if not ai_initialized:
            print("\n  ⚠️  Running without AI. Get free key at: https://console.groq.com")
    else:
        print("\n  ⚠️  Groq not installed. Run: pip install groq")

    if EXCEL_ENABLED:       print("  ✅ Excel creation enabled")
    if REQUESTS_ENABLED:    print("  ✅ Weather & News enabled")
    if PSUTIL_ENABLED:      print("  ✅ App closer enabled")
    if BRIGHTNESS_ENABLED:  print("  ✅ Brightness control enabled")

    print("\n  🎯 CHOOSE MODE:")
    print("  1 → Wake/Sleep Mode (Recommended)")
    print("  2 → Simple Mode (always listening)")

    choice = input("\n  Enter 1 or 2: ").strip()
    print("\n" + "=" * 65)

    if choice == "2":
        run_simple_mode()
    else:
        run_nova()
